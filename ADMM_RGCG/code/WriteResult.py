import openpyxl   #用于处理Excel文件
def WriteResult(method,result_path,TightFlag,y_0,ps, Uplift,KPointNum,iframp,M, RCutNum,time_initialization, iterate_num, time_solve,sigma,eps,beta,maxKPointNum,maxRCutNum,eta, y_max, y_min,NumericTrouble,ub,lb,iterate_uplift,UB,G_1,G_2,G_3,DualFlag,sNum):
    ################################################## 打开已经建好的工作簿
    workbook = openpyxl.load_workbook(result_path)  # 打开记录计算结果的工作簿
    price_sheet = workbook['price']  # 工作表保存价格
    Uplift_sheet = workbook['Uplift']  # 工作表保存Uplift
    KPoint_sheet =workbook['KPoint']  # 工作表保存K_i中点的个数
    RCut_sheet =workbook['RCut']  # 工作表保存R_i中割平面的个数
    G_sheet = workbook['G']  # 工作表保存每个机组集中的机组索引号
    time_sheet = workbook['time']  # 工作表保存求解时间
    iterate_sheet = workbook['iterate']  # 工作表保存迭代次数
    UB_sheet = workbook['UB']  # 工作表保存每次迭代得到的上界UB
    parameter_sheet = workbook['parameter']  # 工作表保存算法参数

    ################################################## 写入数据到工作表
    if method == 'NCG':
        if DualFlag:
            Pc = 3
            j = 3
            g = 4
        else:
            Pc = 2
            j = 0
            g = 0
    elif method == 'LM':
        Pc = 4
        j = 6

    ################################################## 写入参数
    if TightFlag != 3:
        parameter_sheet.cell(2, Pc, sigma)  # 记录算法参数
        parameter_sheet.cell(3, Pc, eps)  # 记录算法参数
        parameter_sheet.cell(4, Pc, beta)  # 记录算法参数
        parameter_sheet.cell(5, Pc, maxKPointNum)  # 记录算法参数
        parameter_sheet.cell(6, Pc, maxRCutNum)  # 记录算法参数
        parameter_sheet.cell(7, Pc, eta)  # 记录算法参数
    for i in range(y_min.size):
        price_sheet.cell(4+i, 21, y_min[i])  # 记录算法参数
        price_sheet.cell(4+i, 22, y_max[i])  # 记录算法参数

    ################################################## 写入Uplift的行标
    Uplift_sheet.cell(3, 1, 'Uplift')  # 记录Uplift
    Uplift_sheet.cell(4, 1, 'NumericTrouble')  # 记录是否存在数值问题
    Uplift_sheet.cell(5, 1, 'ub')  # 记录对偶问题函数值上界
    Uplift_sheet.cell(6, 1, 'lb')  # 记录对偶问题函数值下界

    if TightFlag == 0:  # 对比实验，用很松的'2Pwan'模型代替凸包模型解UC松弛问题，得到松弛解x_0
        for i in range(len(y_0)):
            price_sheet.cell(i + 4, 1, i + 1)  # 记录时段
            price_sheet.cell(i + 4, 3+j*2, y_0[i])  # 记录2P模型的初始价格
            price_sheet.cell(i + 4, 4+j*2, ps[i])  # 记录2P模型的最终价格
        Uplift_sheet.cell(3, 3+j, Uplift)  # 记录2P模型的Uplift
        Uplift_sheet.cell(4, 3+j, NumericTrouble)  # 记录是否存在数值问题
        Uplift_sheet.cell(5, 3 + j, ub)  # 记录对偶问题函数值上界
        Uplift_sheet.cell(6, 3 + j, lb)  # 记录对偶问题函数值下界
        for i in range(len(iterate_uplift)):   # 记录每次迭代得到的uplift
            Uplift_sheet.cell(i + 7, 1, i + 1)  # 记录迭代第几次
            Uplift_sheet.cell(i + 7, 3+j, iterate_uplift[i])  # 记录2P模型本次迭代得到的uplift
        time_sheet.cell(3, 3+j, time_initialization)  # 记录2P模型求解初始价格的时间
        time_sheet.cell(4, 3 + j, time_solve)  # 记录用2P模型整个算法的求解时间
        time_sheet.cell(5, 3+j, time_initialization+time_solve)  # 记录用2P模型整个算法的总时间
        iterate_sheet.cell(3, 2 + j, iterate_num)  # 记录2P模型迭代次数
        for i in range(len(UB)):
            UB_sheet.cell(i + 3, 1, i + 1)  # 记录迭代第几次
            UB_sheet.cell(i + 3, 2+j, UB[i])  # 记录2P模型本次迭代得到的UB
    elif TightFlag == 1:  # 用'TP'模型代替凸包模型解UC松弛问题，得到松弛解x_0
        for i in range(len(y_0)):
            price_sheet.cell(i + 4, 5+j*2, y_0[i])  # 记录TP模型的初始价格
            price_sheet.cell(i + 4, 6+j*2, ps[i])  # 记录TP模型的最终价格
        Uplift_sheet.cell(3, 4+j, Uplift)  # 记录TP模型的Uplift
        Uplift_sheet.cell(4, 4 + j, NumericTrouble)  # 记录是否存在数值问题
        Uplift_sheet.cell(5, 4 + j, ub)  # 记录对偶问题函数值上界
        Uplift_sheet.cell(6, 4 + j, lb)  # 记录对偶问题函数值下界
        for i in range(len(iterate_uplift)):   # 记录每次迭代得到的uplift
            Uplift_sheet.cell(i + 7, 1, i + 1)  # 记录迭代第几次
            Uplift_sheet.cell(i + 7, 4+j, iterate_uplift[i])  # 记录2P模型本次迭代得到的uplift
        time_sheet.cell(3, 4+j, time_initialization)  # 记录TP模型求解初始价格的时间
        time_sheet.cell(4, 4+j, time_solve)  # 记录用TP模型整个算法的求解时间
        time_sheet.cell(5, 4 + j, time_initialization + time_solve)  # 记录用TP模型整个算法的总时间
        iterate_sheet.cell(3, 3 + j, iterate_num)  # 记录TP模型迭代次数
        for i in range(len(UB)):
            UB_sheet.cell(i + 3, 1, i + 1)  # 记录迭代第几次
            UB_sheet.cell(i + 3, 3+j, UB[i])  # 记录TP模型本次迭代得到的UB
    elif TightFlag == 2:  # 初始价格随机
        for i in range(len(y_0)):
            price_sheet.cell(i + 4, 7+j*2, y_0[i])  # 记录TP模型的初始价格
            price_sheet.cell(i + 4, 8+j*2, ps[i])  # 记录TP模型的最终价格
        Uplift_sheet.cell(3, 5+j, Uplift)  # 记录Uplift
        Uplift_sheet.cell(4, 5 + j, NumericTrouble)  # 记录是否存在数值问题
        Uplift_sheet.cell(5, 5 + j, ub)  # 记录对偶问题函数值上界
        Uplift_sheet.cell(6, 5 + j, lb)  # 记录对偶问题函数值下界
        for i in range(len(iterate_uplift)):   # 记录每次迭代得到的uplift
            Uplift_sheet.cell(i + 7, 1, i + 1)  # 记录迭代第几次
            Uplift_sheet.cell(i + 7, 5+j, iterate_uplift[i])  # 记录2P模型本次迭代得到的uplift
        time_sheet.cell(3, 5+j, time_initialization)  # 记录求解初始价格的时间
        time_sheet.cell(4, 5+j, time_solve)  # 记录用整个算法的求解时间
        time_sheet.cell(5, 5 + j, time_initialization + time_solve)  # 记录用2P模型整个算法的总时间
        iterate_sheet.cell(3, 4 + j, iterate_num)  # 记录迭代次数
        for i in range(len(UB)):
            UB_sheet.cell(i + 3, 1, i + 1)  # 记录迭代第几次
            UB_sheet.cell(i + 3, 4+j, UB[i])  # 记录TP模型本次迭代得到的UB
    elif TightFlag == 3:  # 直接用凸包模型
        for i in range(len(y_0)):
            price_sheet.cell(i + 4, 2, ps[i])  # 记录凸包价格
        Uplift_sheet.cell(3, 2, Uplift)  # 记录TP模型的Uplift
        Uplift_sheet.cell(4, 2, NumericTrouble)  # 记录是否存在数值问题
        Uplift_sheet.cell(5, 2, ub)  # 记录对偶问题函数值上界
        Uplift_sheet.cell(6, 2, lb)  # 记录对偶问题函数值下界
        for i in range(len(iterate_uplift)):   # 记录每次迭代得到的uplift
            Uplift_sheet.cell(i + 7, 1, i + 1)  # 记录迭代第几次
            Uplift_sheet.cell(i + 7, 2, iterate_uplift[i])  # 记录2P模型本次迭代得到的uplift
        time_sheet.cell(3, 2, time_initialization)  # 记录TP模型求解初始价格的时间
        time_sheet.cell(4, 2, time_solve)  # 记录用TP模型整个算法的求解时间
        time_sheet.cell(5, 2, time_initialization + time_solve)  # 记录用TP模型整个算法的总时间

    ################################################## 写入KPoint
    if method == 'NCG' and TightFlag != 3:
        for i in range(len(KPointNum)):
            KPoint_sheet.cell(i + 3, 1, i + 1)  # 记录机组编号
            KPoint_sheet.cell(i + 3, 2+TightFlag+j, KPointNum[i])  # 记录模型K_i中添加点的个数
            KPoint_sheet.cell(i + 3, 8, iframp[i])  # 记录模型机组是否考虑爬坡
            KPoint_sheet.cell(i + 3, 9, M[i])  # 记录模型机组的爬坡到最大功率的时段数M

    ################################################## 写入RCut
    if method == 'NCG' and TightFlag != 3:
        for i in range(len(RCutNum)):
            RCut_sheet.cell(i + 3, 1, i + 1)  # 记录机组编号
            RCut_sheet.cell(i + 3, 2 + TightFlag+j, RCutNum[i])  # 记录模型R_i中添加点的个数
            RCut_sheet.cell(i + 3, 8, iframp[i])  # 记录模型机组是否考虑爬坡
            RCut_sheet.cell(i + 3, 9, M[i])  # 记录模型机组的爬坡到最大功率的时段数M
            RCut_sheet.cell(i + 3, 10 + TightFlag, sNum[i])  # 记录每个机组不需要求子问题的次数

    ################################################## 写入G
    for idx, item in enumerate(G_1, start=2):
        G_sheet.cell(idx, 1+g, item+1)  # 记录机组集G1中的机组索引号
    for idx, item in enumerate(G_2, start=2):
        G_sheet.cell(idx, 2+g, item+1)  # 记录机组集G2中的机组索引号
    for idx, item in enumerate(G_3, start=2):
        G_sheet.cell(idx, 3+g, item+1)  # 记录机组集G3中的机组索引号
    # for idx, item in enumerate(G_4, start=2):
    #     G_sheet.cell(idx, 4+g, item+1)  # 记录机组集G4中的机组索引号

    workbook.save(result_path)  # 保存工作簿到文件夹result_UC_AF，文件名为算例名